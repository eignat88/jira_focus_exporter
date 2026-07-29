"""Export ETL results to Excel."""

import os
from datetime import datetime
from typing import List, Dict, Any, Optional


def export_to_excel(
    data: List[Dict[str, Any]],
    filename: str,
    sheet_name: str = "ETL Results",
    output_dir: str = "reports",
) -> str:
    """
    Export data to Excel file.

    Args:
        data: List of dictionaries to export
        filename: Output filename
        sheet_name: Sheet name
        output_dir: Output directory

    Returns:
        Path to generated file
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        # Fallback to CSV if openpyxl not available
        return export_to_csv(data, filename, output_dir)

    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    if not data:
        wb.save(filepath)
        return filepath

    # Headers
    headers = list(data[0].keys())
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Data rows
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, header in enumerate(headers, 1):
            value = row_data.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

    # Auto-width columns
    for col in range(1, len(headers) + 1):
        max_length = max(
            len(str(ws.cell(row=r, column=col).value or ""))
            for r in range(1, len(data) + 2)
        )
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = max_length + 2

    wb.save(filepath)
    return filepath


def export_to_csv(
    data: List[Dict[str, Any]],
    filename: str,
    output_dir: str = "reports",
) -> str:
    """Export data to CSV file."""
    import csv

    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, filename.replace('.xlsx', '.csv'))

    if not data:
        return filepath

    headers = list(data[0].keys())

    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(data)

    return filepath


def export_load_history(
    history: List[Dict[str, Any]],
    output_dir: str = "reports",
) -> str:
    """Export load history to Excel."""
    filename = f"load_history_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return export_to_excel(history, filename, sheet_name="Load History", output_dir=output_dir)


def export_chunk_details(
    chunks: List[Dict[str, Any]],
    output_dir: str = "reports",
) -> str:
    """Export chunk details to Excel."""
    filename = f"chunk_details_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return export_to_excel(chunks, filename, sheet_name="Chunk Details", output_dir=output_dir)


def export_quality_checks(
    checks: List[Dict[str, Any]],
    output_dir: str = "reports",
) -> str:
    """Export quality checks to Excel."""
    filename = f"quality_checks_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return export_to_excel(checks, filename, sheet_name="Quality Checks", output_dir=output_dir)
